import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import tkinter as tk
from tkinter import simpledialog
from config import EXCEL_FAJLOK, OUTPUT_MAPPAK

# Fajl eleresi utja - config fajlban van
EXCEL_PATH = EXCEL_FAJLOK["EXCEL_DATA1"]

def calculate_cost(item_id):
    print(f"Adatok betoltese...")
    
    # Lapok beolvasasa
    try:
        main_data = pd.read_excel(EXCEL_PATH, sheet_name="MAIN_DATA")
        item_list = pd.read_excel(EXCEL_PATH, sheet_name="ITEM_LIST")
        price_list = pd.read_excel(EXCEL_PATH, sheet_name="PRICE_LIST")
    except FileNotFoundError:
        print(f"HIBA: Excel fajl nem talalhato: {EXCEL_PATH}")
        return
    except Exception as e:
        print(f"HIBA: Excel beolvasas sikertelen: {e}")
        return

    # Szures ID-ra
    filtered = main_data[main_data["JOB_ID"].astype(str) == str(item_id)].copy()
    
    if filtered.empty:
        print(f"Nem talalhato ID: {item_id}")
        return

    print(f"Talalt sorok: {len(filtered)} db")

    # ID-nkent osszesiti a fogyasi aranyt
    aggregated = filtered.groupby("ITEM_CODE")["USAGE_RATIO"].sum().reset_index()
    aggregated.columns = ["ITEM_CODE", "Total_usage"]

    # Join: item_list (ITEM_CODE -> PART_NUMBER + szorzo)
    aggregated = aggregated.merge(
        item_list[["CODE", "PART_NUMBER", "COST_MULTIPLIER"]],
        left_on="ITEM_CODE",
        right_on="CODE",
        how="left"
    )

    # Join: price_list (PART_NUMBER -> AVG_PRICE)
    aggregated = aggregated.merge(
        price_list[["PART_CODE", "AVG_PRICE"]],
        left_on="PART_NUMBER",
        right_on="PART_CODE",
        how="left"
    )

    # Koltseg kiszamitasa
    aggregated["Cost_EUR"] = (
        aggregated["Total_usage"] * 
        aggregated["COST_MULTIPLIER"] * 
        aggregated["AVG_PRICE"]
    )

    # NaN ellenőrzés - hiányzó item_list egyezések
    nan_item = aggregated[aggregated["PART_NUMBER"].isna()]
    if not nan_item.empty:
        print(f"FIGYELEM: {len(nan_item)} elem nem talalhato az item listaban:")
        for _, row in nan_item.iterrows():
            print(f"  - ITEM_CODE: {row['ITEM_CODE']}")

    # NaN ellenőrzés - hiányzó ár egyezések
    nan_price = aggregated[aggregated["AVG_PRICE"].isna() & aggregated["PART_NUMBER"].notna()]
    if not nan_price.empty:
        print(f"FIGYELEM: {len(nan_price)} elemnek nincs ara az arlistaban:")
        for _, row in nan_price.iterrows():
            print(f"  - PART_NUMBER: {row['PART_NUMBER']}")

    # Összesített figyelmeztetés
    nan_cost = aggregated["Cost_EUR"].isna().sum()
    if nan_cost > 0:
        print(f"FIGYELEM: {nan_cost} sor koltsege nem szamithato - ezek kiesnek az osszesboll!")

    # Csak a relevas oszlopok
    output_df = aggregated[[
        "ITEM_CODE",
        "PART_NUMBER", 
        "Total_usage",
        "COST_MULTIPLIER",
        "AVG_PRICE",
        "Cost_EUR"
    ]].copy()

    # Osszesen sor
    total_row = pd.DataFrame([{
        "ITEM_CODE": "TOTAL",
        "PART_NUMBER": "",
        "Total_usage": "",
        "COST_MULTIPLIER": "",
        "AVG_PRICE": "",
        "Cost_EUR": aggregated["Cost_EUR"].sum()
    }])
    
    output_df = pd.concat([output_df, total_row], ignore_index=True)

    # Kimenet fajl neve
    output_path = os.path.join(OUTPUT_MAPPAK["KOLTSEG"], f"{item_id}_cost.xlsx")

    # Mentes Excelbe
    try:
        output_df.to_excel(output_path, index=False, sheet_name="Cost_calculation")
    except PermissionError:
        print(f"HIBA: A fajl nyitva van Excelben, zard be es probald ujra!")
        return
    except Exception as e:
        print(f"HIBA: Mentes sikertelen: {e}")
        return

    # Formázás
    wb = load_workbook(output_path)
    ws = wb.active

    # Fejlec formázas (kek hatter, feher bold)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Osszesen sor formázas (sarga hatter, bold)
    last_row = ws.max_row
    summary_fill = PatternFill("solid", fgColor="FFD700")
    summary_font = Font(bold=True)
    
    for cell in ws[last_row]:
        cell.fill = summary_fill
        cell.font = summary_font

    # Oszlopszelesseg auto
    for col in ws.columns:
        max_width = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = max_width

    # EUR formatum a Cost oszlopra
    for cell in ws[f"F2:F{last_row}"]:
        for c in cell:
            c.number_format = '#,##0.00 "EUR"'

    wb.save(output_path)
    print(f"Kesz! Fajl mentve: {output_path}")
    print(f"Teljes koltseg: {aggregated['Cost_EUR'].sum():.2f} EUR")

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    item_id = simpledialog.askstring(
        "ID", 
        "Írja be az azonositot:"
    )
    if item_id:
        calculate_cost(item_id)
